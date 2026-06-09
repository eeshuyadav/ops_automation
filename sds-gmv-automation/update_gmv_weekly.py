"""
Weekly GMV Report updater (Google Sheets edition).

Reads a one-week SDS workflow xlsx, aggregates successful-payment TXN AMT per
merchant, and appends a new week column to the SDS_GMV WEEKLY tab of a live
Google Sheet.

Setup (one-time):
    1. pip install gspread google-auth openpyxl
    2. Create a GCP project, enable the Google Sheets API.
    3. Create a service account, download its JSON key, save next to this
       script as `credentials.json` (or pass --creds <path>).
    4. Share the target Google Sheet with the service account's email as
       Editor (see the `client_email` field in credentials.json).

Usage:
    python update_gmv_weekly.py <sds_workflow.xlsx> \\
        [--sheet-id <id>] [--gid <gid>] [--creds <credentials.json>] \\
        [--dry-run]

Defaults target the GMV spreadsheet provided by ops:
    https://docs.google.com/spreadsheets/d/1-SXDB4Sem3aZaTwWSnUbQtnnS-wJCuzX/edit?gid=900236190
"""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

try:
    import gspread
    from google.oauth2.credentials import Credentials  # OAuth user credentials
    from google.auth.transport.requests import Request
except ImportError:  # pragma: no cover
    raise SystemExit(
        "Missing deps. Run:  pip install gspread google-auth google-auth-oauthlib openpyxl"
    )


GMV_TAB_NAME = "SDS_GMV WEEKLY"
SUCCESS_STATUS = "payment_successful"

# Merchants the QuickSight dashboard's "EXCLUDE MERCHANTS" control filters
# out. Source-of-truth is ops; update this set if the list changes.
# Confirmed from the Ops Dashboard V3 → SALE TXNs tab on 2026-04-21.
EXCLUDED_MERCHANTS: frozenset[int] = frozenset({6792, 3500, 3742, 2928, 13947, 9462})

# Live production GMV sheet that this script writes to. Authenticated as
# chargeback-automation@gokwik.co (OAuth user, see get_oauth_creds below).
DEFAULT_SPREADSHEET_ID = "1UhGckgb4OYauZxJCq4sw3WR2XVWwJ_ew-KkzjdGgndY"
DEFAULT_GID = 0

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# Canonical column names in the source export (order-independent lookup by header)
H_MERCHANT_ID = "MERCHANT ID"
H_ORDER_DATE = "ORDER DATE"
H_BANK_STATUS = "BANK STATUS"
H_TXN_AMT = "TXN AMT"


def ordinal_suffix(day: int) -> str:
    if 10 <= day % 100 <= 20:
        return "th"
    return {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")


def format_week_header(start: datetime, end: datetime) -> str:
    """Format like '6th April - 12th April 26' (end date's 2-digit year)."""
    s = f"{start.day}{ordinal_suffix(start.day)} {start.strftime('%B')}"
    e = f"{end.day}{ordinal_suffix(end.day)} {end.strftime('%B')}"
    return f"{s} - {e} {end.strftime('%y')}"


def col_letter(col: int) -> str:
    """1 -> A, 27 -> AA, etc."""
    out = ""
    while col:
        col, rem = divmod(col - 1, 26)
        out = chr(65 + rem) + out
    return out


_DATE_FORMATS = (
    "%b %d, %Y",              # "Apr 13, 2026"              (xlsx string export)
    "%d-%m-%Y %H:%M",         # "13-04-2026 00:00"          (csv export)
    "%d-%m-%Y",               # "13-04-2026"
    "%Y-%m-%d %H:%M:%S",      # "2026-04-13 00:00:00"
    "%Y-%m-%d",               # "2026-04-13"
    "%Y-%m-%dT%H:%M:%S.%f",   # "2026-04-13T05:49:30.053"   (Metabase MBQL, Z stripped)
    "%Y-%m-%dT%H:%M:%S",      # "2026-04-13T05:49:30"       (Metabase MBQL, no millis)
)


def _parse_date(val) -> datetime | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if s.endswith("Z"):
        s = s[:-1]  # fromisoformat / strptime don't grok trailing Z pre-3.11
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _header_indices(header_row: list) -> dict[str, int]:
    """Map canonical header name → 0-based index. Strips BOM and whitespace."""
    out: dict[str, int] = {}
    for i, h in enumerate(header_row):
        if h is None:
            continue
        key = str(h).lstrip("\ufeff").strip()
        out[key] = i
    return out


def _iter_source_rows(sds_path: Path):
    """Yield (header_indices, row_values_list) for each data row of xlsx or csv."""
    suffix = sds_path.suffix.lower()
    if suffix == ".csv":
        with open(sds_path, "r", newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if header is None:
                raise SystemExit(f"CSV has no header: {sds_path.name}")
            idx = _header_indices(header)
            _require_cols(idx, sds_path)
            for row in reader:
                yield idx, row
        return

    if suffix in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(sds_path, data_only=True, read_only=True)
        # Find a sheet whose first ~20 rows contain the MERCHANT ID header.
        target_ws = None
        header_row_idx = 0
        idx: dict[str, int] = {}
        for ws in wb.worksheets:
            for r_num, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
                if row and any(str(c).lstrip("\ufeff").strip() == H_MERCHANT_ID for c in row if c is not None):
                    target_ws = ws
                    header_row_idx = r_num
                    idx = _header_indices(list(row))
                    break
            if target_ws is not None:
                break
        if target_ws is None:
            raise SystemExit(f"No sheet with '{H_MERCHANT_ID}' header found in {sds_path.name}")
        _require_cols(idx, sds_path)
        for row in target_ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            yield idx, list(row)
        wb.close()
        return

    raise SystemExit(f"Unsupported file type: {suffix}  (use .xlsx or .csv)")


def _require_cols(idx: dict[str, int], path: Path) -> None:
    missing = [h for h in (H_MERCHANT_ID, H_ORDER_DATE, H_BANK_STATUS, H_TXN_AMT) if h not in idx]
    if missing:
        raise SystemExit(f"{path.name} is missing required columns: {missing}")


def aggregate_sds(sds_path: Path) -> tuple[dict[int, float], datetime, datetime]:
    """Return (merchant_id -> sum(TXN AMT), min_date, max_date) for successful txns.
    Accepts QuickSight xlsx or csv export."""
    sums: dict[int, float] = defaultdict(float)
    min_date: datetime | None = None
    max_date: datetime | None = None
    rows_seen = 0

    for idx, row in _iter_source_rows(sds_path):
        if not row or idx[H_MERCHANT_ID] >= len(row):
            continue
        if row[idx[H_BANK_STATUS]] != SUCCESS_STATUS:
            continue
        rows_seen += 1
        mid_raw = row[idx[H_MERCHANT_ID]]
        amt_raw = row[idx[H_TXN_AMT]]
        date_raw = row[idx[H_ORDER_DATE]]
        if mid_raw in (None, "") or amt_raw in (None, ""):
            continue
        try:
            sums[int(float(str(mid_raw).strip()))] += float(str(amt_raw).replace(",", "").strip())
        except (TypeError, ValueError):
            continue
        dt = _parse_date(date_raw)
        if dt is not None:
            if min_date is None or dt < min_date:
                min_date = dt
            if max_date is None or dt > max_date:
                max_date = dt

    if min_date is None or max_date is None:
        raise SystemExit("No valid ORDER DATE found among successful transactions.")
    return dict(sums), min_date, max_date


def _is_numeric_str(s: str) -> bool:
    if s is None or s == "":
        return False
    try:
        float(s.replace(",", "").replace("\u20b9", "").strip())
        return True
    except (TypeError, ValueError):
        return False


def _latest_non_numeric(row_values: list[str], upto_col: int) -> str | None:
    """Walk columns upto_col..3 in a 1-indexed row; return most recent non-numeric, non-empty value.

    Skips Sheets error strings ('#VALUE!', '#REF!', '#N/A', etc.) — those bubble
    up from formula columns (e.g. a 'Difference' column = CP-CQ) and must not be
    carried forward as if they were operational markers.
    """
    # row_values is 0-indexed; cols 1 and 2 are Merchant ID / Name — skip them.
    for c in range(upto_col, 2, -1):
        if c - 1 >= len(row_values):
            continue
        v = row_values[c - 1]
        if v is None or v == "":
            continue
        if _is_numeric_str(v):
            continue
        if isinstance(v, str) and v.startswith("#"):
            continue
        return v
    return None


def open_worksheet(sheet_id: str, gid: int, creds_path: Path):
    """Authenticate as the chargeback-automation@gokwik.co OAuth user and
    open the GMV worksheet for read+write. Refreshes the token on disk if
    it expired."""
    creds = Credentials.from_authorized_user_file(str(creds_path), scopes=SCOPES)
    if not creds.valid:
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
            creds_path.write_text(creds.to_json())
        else:
            raise SystemExit(
                f"OAuth token at {creds_path} is invalid and cannot be "
                "refreshed. Re-run the OAuth flow that produced this token."
            )
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)
    for ws in sh.worksheets():
        if ws.id == gid:
            return sh, ws
    titles = [f"{w.title} (gid={w.id})" for w in sh.worksheets()]
    raise SystemExit(
        f"No worksheet with gid={gid} in spreadsheet {sheet_id}. Tabs: {titles}"
    )


def append_week_column(
    worksheet,
    sums: dict[int, float],
    header: str,
    dry_run: bool,
) -> tuple[int, int, int, int, int]:
    """Append a new column to the GMV tab. Returns (written, carried, blank, unmatched_sds, new_col)."""
    all_values: list[list[str]] = worksheet.get_all_values()
    if not all_values:
        raise SystemExit("GMV tab is empty.")

    header_row = all_values[0]
    # Find last populated column in the header row; new column sits just after it.
    last_populated = 0
    for i, v in enumerate(header_row, 1):
        if v not in (None, ""):
            last_populated = i
    new_col = last_populated + 1
    prev_col = last_populated

    written = 0
    carried = 0
    blank = 0
    matched_ids: set[int] = set()
    grand_total_row: int | None = None  # 1-indexed sheet row

    # Build the column values (index 0 = header row, index 1 = first data row, etc.)
    column: list[str] = [header]
    for r_idx in range(1, len(all_values)):
        row_values = all_values[r_idx]
        mid_raw = row_values[0] if row_values else ""
        if mid_raw in (None, ""):
            column.append("")
            continue

        if str(mid_raw).strip().lower() == "grand total":
            grand_total_row = r_idx + 1  # placeholder; filled after loop
            column.append("")
            continue

        try:
            mid = int(float(str(mid_raw).replace(",", "").strip()))
        except (TypeError, ValueError):
            column.append("")
            continue

        # Note: EXCLUDED_MERCHANTS are filtered upstream in the MBQL fetch,
        # so they can never appear in `sums` here. We deliberately do NOT add
        # an explicit check for them — letting them fall through to the
        # marker-carry-forward path preserves the operational marker
        # (e.g. 'SDS disabled') that ops put in earlier columns. An explicit
        # `column.append("")` here would silently erase that marker.

        if mid in sums:
            column.append(f"{round(sums[mid], 2)}")
            matched_ids.add(mid)
            written += 1
        else:
            marker = _latest_non_numeric(row_values, prev_col)
            if marker is not None:
                column.append(marker)
                carried += 1
            else:
                column.append("")
                blank += 1

    if grand_total_row is not None:
        # Last data row is the merchant row right before Grand Total.
        last_data_row = grand_total_row - 1
        column[grand_total_row - 1] = f"=SUM({col_letter(new_col)}2:{col_letter(new_col)}{last_data_row})"

    unmatched_sds = len(set(sums.keys()) - matched_ids)

    if dry_run:
        return written, carried, blank, unmatched_sds, new_col

    cl = col_letter(new_col)
    rng = f"{cl}1:{cl}{len(column)}"

    # Sheets has a per-tab column cap; if our new column lands past the right
    # edge (e.g. after the user manually added 'Manual'/'Difference' columns),
    # expand the grid first or the update call fails with a 400.
    if worksheet.col_count < new_col:
        worksheet.add_cols(new_col - worksheet.col_count)
    if worksheet.row_count < len(column):
        worksheet.add_rows(len(column) - worksheet.row_count)

    worksheet.update(
        range_name=rng,
        values=[[v] for v in column],
        value_input_option="USER_ENTERED",
    )

    # Copy formatting from the previous week column onto the new one so the
    # Grand Total row stays highlighted and numbers render consistently.
    if prev_col >= 3:
        worksheet.spreadsheet.batch_update({
            "requests": [{
                "copyPaste": {
                    "source": {
                        "sheetId": worksheet.id,
                        "startRowIndex": 0, "endRowIndex": len(column),
                        "startColumnIndex": prev_col - 1, "endColumnIndex": prev_col,
                    },
                    "destination": {
                        "sheetId": worksheet.id,
                        "startRowIndex": 0, "endRowIndex": len(column),
                        "startColumnIndex": new_col - 1, "endColumnIndex": new_col,
                    },
                    "pasteType": "PASTE_FORMAT",
                }
            }]
        })

    return written, carried, blank, unmatched_sds, new_col


_ENABLE_DATE_FORMATS = (
    "%d-%B-%Y", "%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d",
    "%d %B %Y", "%d %b %Y", "%d/%m/%Y", "%d.%m.%Y",
)


def _parse_enable_date(s: str):
    """Parse 'Date of Enabling' from MID tracker (various formats). None if
    unparseable. Distinct from the CSV ORDER DATE parser (_parse_date) — that
    handles ISO 8601 and the manual-export's '%d-%m-%Y %H:%M' shapes."""
    import datetime as _dt, re as _re
    if not s: return None
    s = _re.sub(r'(\d)(st|nd|rd|th)\b', r'\1', s.strip(), flags=_re.IGNORECASE)
    for fmt in _ENABLE_DATE_FORMATS:
        try: return _dt.datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None


def get_tracker_merchants(
    sheet_id: str,
    gid: int,
    oauth_token_path: Path,
    week_end=None,
) -> dict[int, str]:
    """Read MID + Merchant Name from SDS MID tracker (cols A, B, C).
    Returns {merchant_id: merchant_name} for merchants where
    Date of Enabling ≤ week_end (so we don't auto-insert rows for merchants
    enabled AFTER the fetch window — those have no SDS-eligible txns yet).
    """
    from googleapiclient.discovery import build
    creds = Credentials.from_authorized_user_file(str(oauth_token_path), scopes=SCOPES)
    if not creds.valid and creds.expired and creds.refresh_token:
        creds.refresh(Request())
        oauth_token_path.write_text(creds.to_json())
    svc = build("sheets", "v4", credentials=creds)

    meta = svc.spreadsheets().get(spreadsheetId=sheet_id).execute()
    tab = next(
        (s["properties"]["title"] for s in meta.get("sheets", [])
         if s["properties"]["sheetId"] == gid),
        None,
    )
    if tab is None:
        raise SystemExit(f"No tab with gid={gid} in tracker {sheet_id}")

    result = svc.spreadsheets().values().get(
        spreadsheetId=sheet_id, range=f"'{tab}'!A:C"
    ).execute()
    out: dict[int, str] = {}
    excluded_by_date: list[tuple[int, str, str]] = []
    for row in result.get("values", [])[1:]:
        if not row:
            continue
        try:
            mid = int(float(str(row[0]).replace(",", "").strip()))
        except (ValueError, TypeError):
            continue
        if mid in EXCLUDED_MERCHANTS:
            continue
        if week_end is not None and len(row) > 2:
            enable_date = _parse_enable_date(row[2] or "")
            if enable_date is not None and enable_date > week_end:
                excluded_by_date.append((mid, row[1] if len(row) > 1 else "", row[2]))
                continue
        name = row[1].strip() if len(row) > 1 and row[1] else ""
        out[mid] = name
    if excluded_by_date:
        print(f"  Tracker date-filter excluded {len(excluded_by_date)} merchants enabled AFTER {week_end}:")
        for mid, name, date in excluded_by_date:
            print(f"    {mid:>8}  {name!r:<25s}  enabled {date!r}")
    return out


def insert_missing_tracker_rows(worksheet, missing: dict[int, str]) -> int:
    """For each tracker merchant not yet in the GMV sheet, insert a new row
    just above 'Grand Total' with: [merchant_id, merchant_name, 'NA'×past_cols].

    Uses `insertDimension` with `inheritFromBefore=True` so each new row
    inherits the cell formatting (background colour, borders, font) of the
    merchant row immediately above — preserves the visual consistency of the
    sheet. Plain `insert_rows` produces unformatted (white) rows.

    The current-week's value is filled by append_week_column afterwards
    (NA carries forward; numeric value overwrites it for merchants with txns).
    Returns the number of rows inserted."""
    if not missing:
        return 0
    col_a = worksheet.col_values(1)
    grand_total_idx = next(  # 1-indexed
        (i for i, v in enumerate(col_a, 1)
         if v.strip().lower() == "grand total"),
        None,
    )
    if grand_total_idx is None:
        raise SystemExit("Grand Total row not found in sheet — refusing to append blindly.")

    header = worksheet.row_values(1)
    last_pop = sum(1 for h in header if h)
    n = len(missing)

    # 1) Insert n blank rows just above Grand Total, inheriting formatting
    #    from the row above (the last existing merchant row).
    insert_idx_zero = grand_total_idx - 1  # 0-indexed start
    worksheet.spreadsheet.batch_update({
        "requests": [{
            "insertDimension": {
                "range": {
                    "sheetId": worksheet.id,
                    "dimension": "ROWS",
                    "startIndex": insert_idx_zero,
                    "endIndex":   insert_idx_zero + n,
                },
                "inheritFromBefore": True,
            }
        }]
    })

    # 2) Write merchant_id + name + NAs into the newly-inserted rows.
    #    They sit at rows [grand_total_idx, grand_total_idx + n - 1] (1-indexed).
    new_rows = [
        [str(mid), missing[mid]] + ["NA"] * (last_pop - 2)
        for mid in sorted(missing)
    ]
    end_letter = ""
    n_col = last_pop
    while n_col:
        n_col, r = divmod(n_col - 1, 26)
        end_letter = chr(65 + r) + end_letter
    rng = f"A{grand_total_idx}:{end_letter}{grand_total_idx + n - 1}"
    worksheet.update(range_name=rng, values=new_rows, value_input_option="USER_ENTERED")
    return n


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Weekly SDS GMV updater → Google Sheets")
    p.add_argument("sds_xlsx", type=Path, help="Path to SDS_Workflow_automation.xlsx")
    p.add_argument("--sheet-id", default=DEFAULT_SPREADSHEET_ID,
                   help="Google spreadsheet ID (from URL)")
    p.add_argument("--gid", type=int, default=DEFAULT_GID,
                   help="Tab gid (from URL #gid=...)")
    p.add_argument("--creds", type=Path,
                   default=Path(__file__).with_name("oauth_token.json"),
                   help="OAuth user token (chargeback-automation@gokwik.co). "
                        "Defaults to oauth_token.json next to this script.")
    p.add_argument("--tracker-id", default="1CPuJSbd4emdVzfUOpr7cyYE0FBLlqlYE3_OHQtQATcY",
                   help="MID tracker spreadsheet ID for new-merchant auto-insert.")
    p.add_argument("--tracker-gid", type=int, default=759157239,
                   help="MID tracker tab gid (default = 'Same Day Settlements Merchants').")
    p.add_argument("--no-tracker-sync", action="store_true",
                   help="Skip the auto-insert of new tracker merchants. "
                        "Default: insert missing merchants as new rows above Grand Total.")
    p.add_argument("--dry-run", action="store_true",
                   help="Compute and print the planned column without writing")
    return p.parse_args()


def main() -> None:
    args = parse_args()

    if not args.sds_xlsx.exists():
        raise SystemExit(f"SDS file not found: {args.sds_xlsx}")
    if not args.dry_run and not args.creds.exists():
        raise SystemExit(
            f"Credentials not found: {args.creds}\n"
            "See the setup block at the top of this script."
        )

    print(f"Reading SDS workflow:  {args.sds_xlsx}")
    sums, start, end = aggregate_sds(args.sds_xlsx)
    header = format_week_header(start, end)
    print(f"  Week detected:       {start.date()} → {end.date()}")
    print(f"  New column header:   {header}")
    print(f"  Merchants with txns: {len(sums)}")
    print(f"  Grand total TXN AMT: {sum(sums.values()):,.2f}")

    print(f"Target spreadsheet:    https://docs.google.com/spreadsheets/d/{args.sheet_id}/edit?gid={args.gid}")

    if args.dry_run:
        print("(dry-run) Skipping auth; computing against an empty sheet is not possible.")
        print("Run without --dry-run to fetch the tab and preview.")
        return

    _, ws = open_worksheet(args.sheet_id, args.gid, args.creds)
    print(f"Opened tab:            '{ws.title}' (gid={ws.id})")

    # Auto-insert rows for tracker merchants not yet in the GMV sheet.
    # Each new row gets 'NA' for every past week column; this week's value
    # is filled by append_week_column below (NA carries forward if no txns;
    # numeric value overwrites it for merchants with txns).
    if not args.no_tracker_sync:
        tracker = get_tracker_merchants(
            args.tracker_id, args.tracker_gid, args.creds,
            week_end=end.date() if hasattr(end, "date") else end,
        )
        sheet_mids: set[int] = set()
        for v in ws.col_values(1)[1:]:
            if not v or v.strip().lower() == "grand total":
                continue
            try:
                sheet_mids.add(int(float(v.replace(",", "").strip())))
            except (ValueError, TypeError):
                continue
        missing = {m: tracker[m] for m in tracker if m not in sheet_mids}
        if missing:
            n = insert_missing_tracker_rows(ws, missing)
            print(f"  Inserted {n} new merchant row(s) above Grand Total:")
            for mid, name in sorted(missing.items()):
                print(f"    {mid:>8}  {name!r}")
        else:
            print("  No new tracker merchants — sheet rows already in sync.")

    written, carried, blank, unmatched, new_col = append_week_column(
        ws, sums, header, dry_run=False
    )
    print(f"  New column index:    {new_col} ({col_letter(new_col)})")
    print(f"  Rows with TXN sum:   {written}")
    print(f"  Rows carried marker: {carried}")
    print(f"  Rows left blank:     {blank}")
    print(f"  SDS merchants not in GMV (skipped): {unmatched}")
    print("Done.")


if __name__ == "__main__":
    main()
