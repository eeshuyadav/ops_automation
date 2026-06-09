"""Parse the chargeback.xls attachment that ships with every PayU First-Level
Chargeback notification mail. The file is named *.xls but the bytes are
actually .xlsx (Excel 2007+), so openpyxl handles it directly.

Sheet layout (16 cols, header row 1):
    PAYU ID | TRANSACTION ID | TRANSACTION AMOUNT | TRANSACTION DATE
    CUSTOMER FIRST NAME | CUSTOMER LAST NAME | CUSTOMER EMAIL | CUSTOMER PHONE
    PRODUCT INFO | CHARGEBACK DATE | CASE NUMBER | TARGET REPLY DATE
    STATUS | CHARGEBACK TYPE | REASON CODE | COMMENTS

Tabs:
    'New'                  -> active chargeback rows (this is what we process)
    'Pending response'     -> awaiting docs from merchant (skip)
    'Insufficient Document'-> docs received but rejected (skip)
    'Expiring Cases'       -> reply window closing (skip)

Public API:
    parse_attachment(bytes) -> list[dict]
        Each dict represents a row in the 'New' tab with keys:
            payu_id, transaction_id (= Payment id), customer_phone,
            customer_email, transaction_amount, transaction_date,
            chargeback_date, case_number, target_reply_date, reason_code,
            chargeback_type, customer_name (first+last), product_info.
"""
from __future__ import annotations

import io
from typing import Any

try:
    import openpyxl
except ImportError as e:
    raise ImportError(
        "openpyxl is required to parse the PayU attachment. "
        "pip install openpyxl"
    ) from e

# Map from canonical attachment header to our internal key.
_COLUMN_MAP: dict[str, str] = {
    "PAYU ID":             "payu_id",
    "TRANSACTION ID":      "transaction_id",
    "TRANSACTION AMOUNT":  "transaction_amount",
    "TRANSACTION DATE":    "transaction_date",
    "CUSTOMER FIRST NAME": "customer_first_name",
    "CUSTOMER LAST NAME":  "customer_last_name",
    "CUSTOMER EMAIL":      "customer_email",
    "CUSTOMER PHONE":      "customer_phone",
    "PRODUCT INFO":        "product_info",
    "CHARGEBACK DATE":     "chargeback_date",
    "CASE NUMBER":         "case_number",
    "TARGET REPLY DATE":   "target_reply_date",
    "STATUS":              "status",
    "CHARGEBACK TYPE":     "chargeback_type",
    "REASON CODE":         "reason_code",
    "COMMENTS":            "comments",
}

# PayU bundles a thread's cases across four sheets in the same workbook.
# We read all of them; each parsed row carries `_sheet` = the source sheet
# so the tracker (and any inline rendering in the reply body) can show the
# correct status for each row.
_ACTIVE_SHEETS = (
    "New",
    "Pending response",
    "Insufficient Document",
    "Expiring Cases",
)


def _norm_cell(v: Any) -> str:
    """Stringify a cell value while stripping None/whitespace and trailing .0
    on integer-valued floats (PAYU IDs and phones come out as floats from
    openpyxl when the cell type is numeric). Also strips the leading
    apostrophe that PayU uses on the CASE NUMBER cell to force a numeric
    string into text mode in Sheets."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if s.startswith("'"):
        s = s[1:]
    return s


def parse_attachment(data: bytes) -> list[dict]:
    """Parse the PayU chargeback xlsx and return all active case rows
    across every sheet PayU uses (New / Pending response / Insufficient
    Document / Expiring Cases).

    Each returned dict carries `_sheet` = the source sheet name so the
    caller can render or log the correct status per row.

    `data` is the raw attachment payload (already base64-decoded by the
    Gmail attachment fetcher). Returns an empty list if every sheet is
    empty.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    out: list[dict] = []
    for sheet_name in _ACTIVE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        sh = wb[sheet_name]
        rows = list(sh.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        header = [(_norm_cell(c) or "").upper() for c in rows[0]]
        idx: dict[str, int] = {}
        for i, h in enumerate(header):
            key = _COLUMN_MAP.get(h)
            if key:
                idx[key] = i

        for raw_row in rows[1:]:
            if not raw_row or all(c is None or str(c).strip() == "" for c in raw_row):
                continue
            row: dict = {key: _norm_cell(raw_row[i])
                         for key, i in idx.items() if i < len(raw_row)}
            first = row.get("customer_first_name", "")
            last = row.get("customer_last_name", "")
            row["customer_name"] = (first + " " + last).strip()
            row["_sheet"] = sheet_name
            # If STATUS cell was blank, fall back to the sheet name so the
            # tracker still shows something meaningful per row.
            if not row.get("status"):
                row["status"] = sheet_name
            out.append(row)
    return out
